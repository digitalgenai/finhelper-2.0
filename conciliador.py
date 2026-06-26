import os
import tempfile

import pandas as pd


class Conciliador:
    """Classe responsável pela leitura, análise e conciliação de arquivos financeiros."""

    TOLERANCIA = 0.01

    # Mapa de colunas por extensão
    COL_LOC = {".csv": "LOC", ".cnf": "loc_cia", ".xlsx": "Localizador/Cód. Confirmação"}
    COL_LIQ = {".csv": "Liquido", ".cnf": "liquido", ".xlsx": "Total Fornec. (-DF)"}
    COL_PAX = {".csv": "Passageiro", ".cnf": "nome_pax", ".xlsx": "Pax"}

    XLSX_EXTRAS = ["Venda Nº", "Cod. Cliente", "Cod. Emissor", "Markup", "Total Tarifa", "Total Taxas",
                   "Total DU/RAV (Bruta)", "Over Agência", "Total Outras Taxas", "Forma Pgt.",
                   "Form", "Nr. Doc", "Data Venda"]

    # Mapeamento campo a campo: (nome_exibição, colunas_fornecedor, coluna_xlsx)
    # colunas_fornecedor = lista de possíveis nomes (CSV e CNF)
    CAMPO_MAP = [
        ("Tarifa",               ["Tarifa R$", "tarifa_brl"],   "Total Tarifa"),
        ("Taxa",                 ["Taxa", "tx_emb"],            "Total Taxas"),
        ("Total DU/RAV (Bruta)", ["TxDU", "repasse_du"],        "Total DU/RAV (Bruta)"),
        ("Over/Incentivo",       ["Incentivo", "incentivo"],    "Over Agência"),
    ]

    # ── Utilitários ──

    @staticmethod
    def moeda_br(v):
        """Converte string monetária brasileira para float."""
        if pd.isna(v) or str(v).strip() == "":
            return 0.0
        if isinstance(v, (int, float)):
            return round(float(v), 2)
        return round(float(str(v).strip().replace(".", "").replace(",", ".")), 2)

    @staticmethod
    def rotulo(ext: str) -> str:
        return "Wintour" if ext == ".xlsx" else "Fornecedor"

    # ── Leitura ──

    def ler(self, caminho: str):
        """Lê arquivo (.xlsx, .csv ou .cnf) e retorna (DataFrame, extensão)."""
        ext = os.path.splitext(caminho)[1].lower()

        if ext == ".xlsx":
            df = pd.read_excel(caminho, header=5, engine="openpyxl")
        elif ext == ".csv":
            df = pd.read_csv(caminho, sep=";", encoding="latin-1", on_bad_lines="skip")
        elif ext == ".cnf":
            # CNF tem ; no final de cada linha de dados, gerando 1 campo extra
            # index_col=False evita que pandas use a primeira coluna como índice
            df = pd.read_csv(caminho, sep=";", encoding="latin-1", on_bad_lines="skip",
                             index_col=False)
        else:
            raise ValueError(f"Formato {ext} não suportado")

        df.columns = df.columns.map(lambda x: str(x).strip())
        df = df.loc[:, ~df.columns.str.startswith("Unnamed", na=False)].dropna(how="all")

        col_loc = self.COL_LOC[ext]
        col_liq = self.COL_LIQ[ext]

        df[col_loc] = df[col_loc].astype(str).str.strip().str.upper()

        if ext == ".xlsx":
            df[col_liq] = df[col_liq].apply(lambda v: round(float(v), 2) if pd.notna(v) else 0.0)
        else:
            df[col_liq] = df[col_liq].apply(self.moeda_br)

        return df, ext

    # ── Agrupamento ──

    # Colunas extras do CSV/CNF que precisamos para comparação campo a campo
    CSV_EXTRAS = ["Tarifa R$", "Taxa", "TxDU", "Incentivo",
                  "tarifa_brl", "tx_emb", "repasse_du", "incentivo", "comissao", "acrescimos",
                  "fee", "Fee", "emissao", "bilhete"]

    def agrupar(self, df: pd.DataFrame, ext: str) -> dict:
        """Agrupa registros por localizador."""
        cl = self.COL_LOC[ext]
        cq = self.COL_LIQ[ext]
        cp = self.COL_PAX[ext]

        grupos = {}
        for _, r in df.iterrows():
            k = str(r[cl]).strip().upper()
            if not k or k in ("NAN", "NONE", "NAT"):
                continue
            item = {"liquido": r[cq], "pax": str(r.get(cp, "")).strip()}
            if ext == ".xlsx":
                forma_pgt = str(r.get("Forma Pgt.", "")).strip().upper()
                if forma_pgt == "XX":
                    continue
                cod_status = str(r.get("Cód. Status", "")).strip().upper()
                item["is_cf"] = cod_status == "CF"
                for c in self.XLSX_EXTRAS:
                    item[c] = str(r.get(c, "")).strip() if pd.notna(r.get(c)) else ""
            elif ext in (".csv", ".cnf"):
                for c in self.CSV_EXTRAS:
                    item[c] = str(r.get(c, "")).strip() if pd.notna(r.get(c)) else ""
            grupos.setdefault(k, []).append(item)
        return grupos

    # ── Extras XLSX ──

    @staticmethod
    def _extras_xlsx(registros: list) -> dict:
        """Extrai dados extras do primeiro registro xlsx."""
        r = registros[0]
        form   = str(r.get("Form", "")).strip()
        nr_doc = str(r.get("Nr. Doc", "")).strip()
        return {
            "venda": r.get("Venda Nº", ""),
            "cliente": r.get("Cod. Cliente", ""),
            "emissor": r.get("Cod. Emissor", ""),
            "markup": r.get("Markup", ""),
            "tarifa": r.get("Total Tarifa", ""),
            "taxas": r.get("Total Taxas", ""),
            "over_agencia": r.get("Over Agência", ""),
            "forma_pgt": r.get("Forma Pgt.", ""),
            "bilhete": form + nr_doc,
            "du_rav": r.get("Total DU/RAV (Bruta)", ""),
            "outras_taxas": r.get("Total Outras Taxas", ""),
            "data_emissao": pd.Timestamp(r.get("Data Venda")).strftime("%d/%m/%Y") if pd.notna(r.get("Data Venda")) and r.get("Data Venda") != "" else "",
        }

    # ── Comparação campo a campo ──

    @staticmethod
    def _safe_float(v):
        """Converte valor para float sem aplicar lógica de moeda BR."""
        if v is None or str(v).strip() in ("", "nan", "NaN", "None"):
            return 0.0
        try:
            return round(float(v), 2)
        except (ValueError, TypeError):
            return 0.0

    def _comparar_campos(self, registros_csv, registros_xlsx):
        """Compara campos correspondentes entre CSV/CNF e XLSX e retorna descrição da divergência."""
        diferencas = []
        for nome, cols_fornec, col_xlsx in self.CAMPO_MAP:
            # Soma dos valores do CSV/CNF (moeda BR: 1.234,56)
            total_csv = 0.0
            for r in registros_csv:
                # Tenta cada coluna possível até encontrar uma com valor
                val = 0
                for col in cols_fornec:
                    v = r.get(col, "")
                    if v and str(v).strip() not in ("", "0", "0.0", "0,00"):
                        val = v
                        break
                    if not val:
                        val = r.get(col, 0)
                total_csv += self.moeda_br(val)
            total_csv = round(total_csv, 2)

            # Soma dos valores do XLSX (já são float ou string decimal: 1234.56)
            total_xlsx = 0.0
            for r in registros_xlsx:
                total_xlsx += self._safe_float(r.get(col_xlsx, 0))
            total_xlsx = round(total_xlsx, 2)

            dif_val = round(abs(total_csv - total_xlsx), 2)
            if dif_val > self.TOLERANCIA:
                diferencas.append({
                    "campo": nome,
                    "fornec": total_csv,
                    "wintour": total_xlsx,
                    "dif": round(total_csv - total_xlsx, 2)
                })

        if diferencas:
            campos = [d["campo"] for d in diferencas]
            detalhe = " | ".join([f"{d['campo']}: Fornec {d['fornec']:.2f} x Wintour {d['wintour']:.2f} (dif: {d['dif']:.2f})" for d in diferencas])
            return {
                "resumo": ", ".join(campos),
                "detalhe": detalhe
            }
        return {"resumo": "", "detalhe": ""}

    # ── Conciliação ──

    def conciliar(self, g1: dict, g2: dict, lbl1: str, lbl2: str, ext1: str, ext2: str) -> list:
        """Compara dois conjuntos de localizadores e retorna lista de resultados."""
        locs1, locs2 = set(g1), set(g2)
        resultado = []

        def get_extras(loc):
            if ext1 == ".xlsx" and loc in g1:
                return self._extras_xlsx(g1[loc])
            if ext2 == ".xlsx" and loc in g2:
                return self._extras_xlsx(g2[loc])
            return {"venda": "", "cliente": "", "emissor": "", "markup": "", "tarifa": "", "taxas": "",
                    "over_agencia": "", "forma_pgt": ""}

        def get_csv_recs(loc):
            if ext1 in (".csv", ".cnf"):
                return g1.get(loc, [])
            if ext2 in (".csv", ".cnf"):
                return g2.get(loc, [])
            return []

        def safe_float(v):
            try:
                return round(float(v), 2)
            except Exception:
                return 0.0

        _over_defaults = {"over_agencia": "", "incentivo_fornecedor": "", "over_dif": "",
                          "tarifa_fornecedor": "", "tarifa_dif": "",
                          "taxa_fornecedor": "", "taxa_dif": "", "forma_pgt": "", "bilhete": "",
                          "du_rav": "", "outras_taxas": "", "taxa_adm_forn": "", "du_forn": "", "fee_forn": "",
                          "data_emissao": ""}

        # Localizadores presentes em ambos
        for loc in sorted(locs1 & locs2):
            s1 = round(sum(r["liquido"] for r in g1[loc]), 2)
            s2 = round(sum(r["liquido"] for r in g2[loc]), 2)
            dif = round(s1 - s2, 2)
            pax = g1[loc][0]["pax"] or g2[loc][0]["pax"]
            extras = get_extras(loc)

            tarifa = safe_float(extras["tarifa"])
            markup = safe_float(extras["markup"])
            taxas = safe_float(extras["taxas"])
            over_agencia = safe_float(extras.get("over_agencia", 0))
            forma_pgt = str(extras.get("forma_pgt", "")).strip().upper()

            # CF (Conferido): venda revisada manualmente no Wintour
            xlsx_group = g1[loc] if ext1 == ".xlsx" else g2[loc]
            is_cf = any(r.get("is_cf", False) for r in xlsx_group)

            # Esperado depende da forma de pagamento:
            # IV → tarifa + taxas são repassados pelo fornecedor, incluir na fórmula
            # Outros (cartão etc.) → desconsiderar tarifa + taxas
            # Over Agência é sempre subtraído (pago pelo fornecedor à agência, reduz o repasse)
            if forma_pgt == "IV":
                esperado = round((tarifa + taxas) - markup - over_agencia, 2)
            else:
                esperado = round(-over_agencia - markup, 2)

            # Over/Incentivo: Over Agência (XLSX) − Incentivo (CSV)
            incentivo_csv = round(
                sum(self.moeda_br(r.get("Incentivo", "") or r.get("incentivo", ""))
                    for r in get_csv_recs(loc)), 2
            )
            over_dif = round(over_agencia - incentivo_csv, 2)

            # Tarifa: Tarifa XLSX − Tarifa CSV/CNF
            tarifa_forn = round(
                sum(self.moeda_br(r.get("Tarifa R$", "") or r.get("tarifa_brl", ""))
                    for r in get_csv_recs(loc)), 2
            )
            tarifa_dif = round(tarifa - tarifa_forn, 2)

            # Taxa de Embarque: Taxas XLSX − Taxa CSV/CNF
            taxa_forn = round(
                sum(self.moeda_br(r.get("Taxa", "") or r.get("tx_emb", ""))
                    for r in get_csv_recs(loc)), 2
            )
            taxa_dif = round(taxas - taxa_forn, 2)

            # Taxa Adm. Cartão: acrescimos do CSV/CNF (Flytour)
            taxa_adm_forn = round(
                sum(self.moeda_br(r.get("acrescimos", ""))
                    for r in get_csv_recs(loc)), 2
            )

            # Taxa DU: TxDU / repasse_du do CSV/CNF (Flytour)
            du_forn = round(
                sum(self.moeda_br(r.get("TxDU", "") or r.get("repasse_du", ""))
                    for r in get_csv_recs(loc)), 2
            )

            # Fee: coluna fee do CSV/CNF (Flytour)
            fee_forn = round(
                sum(self.moeda_br(r.get("fee", "") or r.get("Fee", ""))
                    for r in get_csv_recs(loc)), 2
            )

            if is_cf:
                status = "Conferido"
            elif abs(dif) < self.TOLERANCIA:
                status = "Ok"
            else:
                status = "Divergente"

            # Comparação campo a campo para identificar origem da divergência
            origem_dif = ""
            origem_dif_detalhe = ""
            if status in ("Divergente", "Conferido") and abs(dif) >= self.TOLERANCIA:
                # Determinar qual grupo é CSV e qual é XLSX
                comp = {"resumo": "", "detalhe": ""}
                if ext1 in (".csv", ".cnf") and ext2 == ".xlsx":
                    comp = self._comparar_campos(g1[loc], g2[loc])
                elif ext1 == ".xlsx" and ext2 in (".csv", ".cnf"):
                    comp = self._comparar_campos(g2[loc], g1[loc])
                origem_dif = comp["resumo"]
                origem_dif_detalhe = comp["detalhe"]

            # Multi-pax: expande em uma linha por Venda Nº do XLSX
            n = len(xlsx_group)

            if n > 1:
                s_csv    = s2 if ext1 == ".xlsx" else s1
                csv_recs = list(get_csv_recs(loc))

                # Pareia registros Wintour com CNF em dois passes:
                # Passo A — bilhete obrigatório: form+nrdoc como sufixo do bilhete CNF.
                # Passo B — pax fallback SOMENTE para CNF sem bilhete (sem bilhete não é possível validar).
                # CNF com bilhete que não casou no passo A → Somente Fornecedor (bilhete não bate).
                # Wintour sem par → Somente Wintour.
                matched_csv  = [False] * len(csv_recs)
                matched_xlsx = [False] * len(xlsx_group)
                xlsx_pairs   = [(rec, None) for rec in xlsx_group]

                def _find_cnf_by_suffix(suffix):
                    for i, cr in enumerate(csv_recs):
                        if matched_csv[i]:
                            continue
                        bil = str(cr.get("bilhete", "")).strip()
                        if suffix and bil.endswith(suffix):
                            return i
                    return None

                # Passo A: bilhete
                for j, rec in enumerate(xlsx_group):
                    form_key   = str(rec.get("Form", "")).strip()
                    nr_doc_key = str(rec.get("Nr. Doc", "")).strip()
                    idx = _find_cnf_by_suffix(form_key + nr_doc_key)
                    if idx is None:
                        idx = _find_cnf_by_suffix(nr_doc_key)
                    if idx is not None:
                        matched_csv[idx]  = True
                        matched_xlsx[j]   = True
                        xlsx_pairs[j]     = (rec, csv_recs[idx])

                # Passo B: fallback por pax — apenas CNF SEM bilhete (não conseguimos validar bilhete)
                csv_no_bil_by_pax = {}
                for i, cr in enumerate(csv_recs):
                    if not matched_csv[i] and not str(cr.get("bilhete", "")).strip():
                        pax_key = str(cr.get("pax", "")).strip().upper()
                        csv_no_bil_by_pax.setdefault(pax_key, []).append(i)

                for j, rec in enumerate(xlsx_group):
                    if matched_xlsx[j]:
                        continue
                    xlsx_pax = str(rec.get("pax", "")).strip().upper()
                    candidates = [i for i in csv_no_bil_by_pax.get(xlsx_pax, []) if not matched_csv[i]]
                    if candidates:
                        idx = candidates[0]
                        csv_no_bil_by_pax[xlsx_pax].remove(idx)
                        matched_csv[idx]  = True
                        matched_xlsx[j]   = True
                        xlsx_pairs[j]     = (rec, csv_recs[idx])

                # CNF que não casaram → Somente Fornecedor (bilhete não correspondeu)
                truly_extra_csv = [cr for i, cr in enumerate(csv_recs) if not matched_csv[i]]

                for rec, cr in xlsx_pairs:
                    ind_liq  = round(rec["liquido"], 2)
                    ind_over = safe_float(rec.get("Over Agência", ""))
                    ind_tar  = safe_float(rec.get("Total Tarifa", ""))
                    ind_tax  = safe_float(rec.get("Total Taxas", ""))
                    form     = str(rec.get("Form", "")).strip()
                    nr_doc   = str(rec.get("Nr. Doc", "")).strip()
                    rec_is_cf = rec.get("is_cf", False)

                    if cr is not None:
                        s_csv_ind    = round(cr["liquido"], 2)
                        tar_forn_ind = round(self.moeda_br(cr.get("Tarifa R$", "") or cr.get("tarifa_brl", "")), 2)
                        tax_forn_ind = round(self.moeda_br(cr.get("Taxa", "") or cr.get("tx_emb", "")), 2)
                        inc_ind      = round(self.moeda_br(cr.get("Incentivo", "") or cr.get("incentivo", "")), 2)
                        taxa_adm_ind = round(self.moeda_br(cr.get("acrescimos", "")), 2)
                        du_ind       = round(self.moeda_br(cr.get("TxDU", "") or cr.get("repasse_du", "")), 2)
                        fee_ind      = round(self.moeda_br(cr.get("fee", "") or cr.get("Fee", "")), 2)
                    else:
                        s_csv_ind    = round(s_csv / n, 2)
                        tar_forn_ind = round(tarifa_forn / n, 2)
                        tax_forn_ind = round(taxa_forn / n, 2)
                        inc_ind      = round(incentivo_csv / n, 2)
                        taxa_adm_ind = round(taxa_adm_forn / n, 2)
                        du_ind       = round(du_forn / n, 2)
                        fee_ind      = round(fee_forn / n, 2)

                    ind_dif = round(ind_liq - s_csv_ind, 2)
                    if cr is None:
                        # Pax do Wintour sem par no CNF — não pode ser Conferido
                        ind_status = "Somente Wintour"
                    elif rec_is_cf:
                        ind_status = "Conferido"
                    elif abs(ind_dif) <= self.TOLERANCIA:
                        ind_status = "Ok"
                    else:
                        ind_status = "Divergente"

                    liq_forn = "" if cr is None else (ind_liq if ext1 == ".xlsx" else s_csv_ind)
                    liq_win  = ind_liq if ext2 == ".xlsx" else s_csv_ind

                    show_origem = ind_status in ("Divergente", "Conferido") and abs(ind_dif) > self.TOLERANCIA
                    resultado.append({
                        "loc": loc,
                        "pax": str(rec.get("pax", "")).strip(),
                        "status": ind_status,
                        f"liq_{lbl1}": liq_forn,
                        f"liq_{lbl2}": liq_win,
                        "dif": ind_dif,
                        "origem_dif": origem_dif if show_origem else "",
                        "origem_dif_detalhe": origem_dif_detalhe if show_origem else "",
                        "over_agencia": ind_over,
                        "incentivo_fornecedor": inc_ind,
                        "over_dif": round(ind_over - inc_ind, 2),
                        "tarifa_fornecedor": tar_forn_ind,
                        "tarifa_dif": round(ind_tar - tar_forn_ind, 2),
                        "taxa_fornecedor": tax_forn_ind,
                        "taxa_dif": round(ind_tax - tax_forn_ind, 2),
                        "taxa_adm_forn": taxa_adm_ind,
                        "du_forn": du_ind,
                        "fee_forn": fee_ind,
                        "forma_pgt": forma_pgt,
                        "venda":    str(rec.get("Venda Nº", "")).strip(),
                        "cliente":  str(rec.get("Cod. Cliente", "")).strip(),
                        "emissor":  str(rec.get("Cod. Emissor", "")).strip(),
                        "markup":   str(rec.get("Markup", "")).strip(),
                        "bilhete":  str(cr.get("bilhete", "")).strip() if cr is not None else form + nr_doc,
                        "du_rav":   str(rec.get("Total DU/RAV (Bruta)", "")).strip(),
                        "outras_taxas": str(rec.get("Total Outras Taxas", "")).strip(),
                        "data_emissao": pd.Timestamp(rec.get("Data Venda")).strftime("%d/%m/%Y") if pd.notna(rec.get("Data Venda")) and rec.get("Data Venda") != "" else "",
                    })
                # Bilhetes extras do CNF sem par no Wintour dentro deste loc
                for extra_cr in truly_extra_csv:
                    extra_liq = round(extra_cr["liquido"], 2)
                    data_em_extra = str(extra_cr.get("emissao", "")).strip()
                    resultado.append({
                        "loc": loc,
                        "pax": str(extra_cr.get("pax", "")).strip(),
                        "status": "Somente Fornecedor",
                        f"liq_{lbl1}": extra_liq if ext1 != ".xlsx" else "",
                        f"liq_{lbl2}": extra_liq if ext2 == ".xlsx" else "",
                        "dif": "",
                        "origem_dif": f"Bilhete extra no {lbl1}",
                        "origem_dif_detalhe": "",
                        **_over_defaults,
                        "incentivo_fornecedor": round(self.moeda_br(extra_cr.get("Incentivo", "") or extra_cr.get("incentivo", "")), 2),
                        "tarifa_fornecedor": round(self.moeda_br(extra_cr.get("Tarifa R$", "") or extra_cr.get("tarifa_brl", "")), 2),
                        "taxa_fornecedor": round(self.moeda_br(extra_cr.get("Taxa", "") or extra_cr.get("tx_emb", "")), 2),
                        "taxa_adm_forn": round(self.moeda_br(extra_cr.get("acrescimos", "")), 2),
                        "du_forn": round(self.moeda_br(extra_cr.get("TxDU", "") or extra_cr.get("repasse_du", "")), 2),
                        "fee_forn": round(self.moeda_br(extra_cr.get("fee", "") or extra_cr.get("Fee", "")), 2),
                        "forma_pgt": forma_pgt,
                        "venda": "", "cliente": "", "emissor": "", "markup": "",
                        "bilhete": str(extra_cr.get("bilhete", "")).strip(), "du_rav": "", "outras_taxas": "",
                        "data_emissao": data_em_extra,
                    })

            else:
                csv_recs_sp = list(get_csv_recs(loc))
                if csv_recs_sp:
                    cnf_bilhete = str(csv_recs_sp[0].get("bilhete", "")).strip()
                    if cnf_bilhete:
                        extras["bilhete"] = cnf_bilhete
                resultado.append({
                    "loc": loc, "pax": pax, "status": status,
                    f"liq_{lbl1}": s1, f"liq_{lbl2}": s2, "dif": dif,
                    "origem_dif": origem_dif,
                    "origem_dif_detalhe": origem_dif_detalhe,
                    "over_agencia": over_agencia,
                    "incentivo_fornecedor": incentivo_csv,
                    "over_dif": over_dif,
                    "tarifa_fornecedor": tarifa_forn,
                    "tarifa_dif": tarifa_dif,
                    "taxa_fornecedor": taxa_forn,
                    "taxa_dif": taxa_dif,
                    "taxa_adm_forn": taxa_adm_forn,
                    "du_forn": du_forn,
                    "fee_forn": fee_forn,
                    "forma_pgt": forma_pgt,
                    **extras,
                })

        def _forn_fields_from_rec(rec):
            """Extrai campos do fornecedor (CNF) para linhas Somente Fornecedor."""
            return {
                "incentivo_fornecedor": round(self.moeda_br(rec.get("Incentivo", "") or rec.get("incentivo", "")), 2),
                "tarifa_fornecedor":    round(self.moeda_br(rec.get("Tarifa R$", "") or rec.get("tarifa_brl", "")), 2),
                "taxa_fornecedor":      round(self.moeda_br(rec.get("Taxa", "") or rec.get("tx_emb", "")), 2),
                "taxa_adm_forn":        round(self.moeda_br(rec.get("acrescimos", "")), 2),
                "du_forn":              round(self.moeda_br(rec.get("TxDU", "") or rec.get("repasse_du", "")), 2),
                "fee_forn":             round(self.moeda_br(rec.get("fee", "") or rec.get("Fee", "")), 2),
            }

        # Somente no grupo 1
        for loc in sorted(locs1 - locs2):
            status = "Somente Fornecedor" if ext1 != ".xlsx" else "Somente Wintour"
            origem = f"Localizador ausente no {lbl2}"
            recs = g1[loc]
            for rec in recs:
                ind_liq = round(rec["liquido"], 2)
                form    = str(rec.get("Form", "")).strip()
                nr_doc  = str(rec.get("Nr. Doc", "")).strip()
                data_em = str(rec.get("Data Venda", "") or rec.get("emissao", "")).strip()
                bilhete_val = str(rec.get("bilhete", "")).strip() if ext1 in (".csv", ".cnf") else form + nr_doc
                venda_val   = "" if ext1 in (".csv", ".cnf") else str(rec.get("Venda Nº", "")).strip()
                extra_forn  = _forn_fields_from_rec(rec) if ext1 in (".csv", ".cnf") else {}
                resultado.append({
                    "loc": loc, "pax": str(rec.get("pax", "")).strip(), "status": status,
                    f"liq_{lbl1}": ind_liq, f"liq_{lbl2}": "", "dif": "",
                    "origem_dif": origem,
                    **_over_defaults,
                    **extra_forn,
                    "bilhete": bilhete_val,
                    "venda":   venda_val,
                    "cliente": str(rec.get("Cod. Cliente", "")).strip(),
                    "emissor": str(rec.get("Cod. Emissor", "")).strip(),
                    "markup":  str(rec.get("Markup", "")).strip(),
                    "data_emissao": data_em,
                })

        # Somente no grupo 2
        for loc in sorted(locs2 - locs1):
            status = "Somente Fornecedor" if ext2 != ".xlsx" else "Somente Wintour"
            origem = f"Localizador ausente no {lbl1}"
            recs = g2[loc]
            for rec in recs:
                ind_liq = round(rec["liquido"], 2)
                form    = str(rec.get("Form", "")).strip()
                nr_doc  = str(rec.get("Nr. Doc", "")).strip()
                data_em = str(rec.get("Data Venda", "") or rec.get("emissao", "")).strip()
                bilhete_val = str(rec.get("bilhete", "")).strip() if ext2 in (".csv", ".cnf") else form + nr_doc
                venda_val   = "" if ext2 in (".csv", ".cnf") else str(rec.get("Venda Nº", "")).strip()
                extra_forn  = _forn_fields_from_rec(rec) if ext2 in (".csv", ".cnf") else {}
                resultado.append({
                    "loc": loc, "pax": str(rec.get("pax", "")).strip(), "status": status,
                    f"liq_{lbl1}": "", f"liq_{lbl2}": ind_liq, "dif": "",
                    "origem_dif": origem,
                    **_over_defaults,
                    **extra_forn,
                    "bilhete": bilhete_val,
                    "venda":   venda_val,
                    "cliente": str(rec.get("Cod. Cliente", "")).strip(),
                    "emissor": str(rec.get("Cod. Emissor", "")).strip(),
                    "markup":  str(rec.get("Markup", "")).strip(),
                    "data_emissao": data_em,
                })

        # INTERFACE detectado em Emissor ou Cliente → Divergente (exceto Somente Wintour/Fornecedor)
        for r in resultado:
            e = str(r.get("emissor", "")).strip().upper()
            c = str(r.get("cliente", "")).strip().upper()
            if e == "EINTERFACE" or c == "CINTERFACE":
                old_status = r["status"]
                if old_status not in ("Somente Wintour", "Somente Fornecedor", "Conferido"):
                    r["status"] = "Divergente"
                # Adicionar EINTERFACE na origem se não tinha outra explicação
                origem = r.get("origem_dif", "")
                if not origem or "ausente" in origem:
                    r["origem_dif"] = "EINTERFACE" + (f" ({old_status})" if old_status not in ("Divergente", "Somente Wintour", "Somente Fornecedor") else "")

        return resultado

    # ── Geração XLSX ──

    def gerar_xlsx(self, resultado: list, lbl1: str, lbl2: str) -> str:
        """Gera planilha Excel estilizada com identidade visual Wee Travel."""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        from collections import Counter

        _STATUS_ORDEM = {"Divergente": 0, "Ok": 1, "Somente Fornecedor": 2, "Conferido": 3}
        registros = [r for r in resultado if r["status"] != "Somente Wintour"]
        registros.sort(key=lambda r: _STATUS_ORDEM.get(r["status"], 99))

        # ── Paleta Wee Travel ──
        COR_HEADER     = "7F2EC2"   # roxo
        COR_RESUMO_BG  = "F3E8FF"   # roxo bem claro
        COR_RESUMO_TTL = "5B21B6"   # roxo escuro

        STATUS_FILL = {
            "Divergente":          "FEE2E2",  # vermelho claro
            "Ok":                  "DCFCE7",  # verde claro
            "Somente Fornecedor":  "CFFAFE",  # ciano claro
            "Conferido":           "EDE9FE",  # lilás
        }

        col_names = [
            "Passageiro", "Cliente", "Emissor", "Origem Dif.", "Detalhe da Diferença",
            "Status", f"Liq. {lbl2}", f"Liq. {lbl1}", "Incentivo (Fornecedor)", "Tarifa Fornecedor",
            "Taxa Embarque", "Taxa DU", "Taxa Adm. Cartão", "Fee", "Markup",
            "Localizador", "Data Emissão", "Nº Venda", "Nº Bilhete",
        ]
        COLS_NUMERICAS = {7, 8, 9, 10, 11, 12, 13, 14}   # 1-based: colunas de valor

        wb = Workbook()
        ws = wb.active
        ws.title = "Conciliação"

        # ── Cabeçalho ──
        hdr_fill  = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(col_names)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(col_names))}1"

        # ── Dados ──
        fnt_data   = Font(name="Calibri", size=10)
        aln_center = Alignment(horizontal="center", vertical="center")
        aln_left   = Alignment(horizontal="left",   vertical="center")

        def _num(v):
            if isinstance(v, (int, float)):
                return round(float(v), 2)
            try:
                return round(float(str(v).replace(",", ".")), 2)
            except (ValueError, TypeError):
                return ""

        for r in registros:
            row_data = [
                r["pax"],
                r.get("cliente", ""),
                r.get("emissor", ""),
                r.get("origem_dif", ""),
                r.get("origem_dif_detalhe", ""),
                r["status"],
                _num(r.get(f"liq_{lbl2}", "")),
                _num(r.get(f"liq_{lbl1}", "")),
                _num(r.get("incentivo_fornecedor", "")),
                _num(r.get("tarifa_fornecedor", "")),
                _num(r.get("taxa_fornecedor", "")),
                _num(r.get("du_forn", "")),
                _num(r.get("taxa_adm_forn", "")),
                _num(r.get("fee_forn", "")),
                r.get("markup", ""),
                r["loc"],
                r.get("data_emissao", ""),
                r.get("venda", ""),
                r.get("bilhete", ""),
            ]
            ws.append(row_data)
            row_n = ws.max_row
            fill = PatternFill(start_color=STATUS_FILL.get(r["status"], "FFFFFF"),
                               end_color=STATUS_FILL.get(r["status"], "FFFFFF"),
                               fill_type="solid")
            for i, cell in enumerate(ws[row_n], 1):
                cell.fill  = fill
                cell.font  = fnt_data
                cell.alignment = aln_center if i in COLS_NUMERICAS else aln_left
                if i in COLS_NUMERICAS and isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'

        # ── Resumo ──
        contagem  = Counter(r["status"] for r in registros)
        total_liq = sum(_num(r.get(f"liq_{lbl1}", 0)) for r in registros
                        if isinstance(_num(r.get(f"liq_{lbl1}", 0)), float))

        ws.append([""] * len(col_names))

        res_fill  = PatternFill(start_color=COR_RESUMO_BG, end_color=COR_RESUMO_BG, fill_type="solid")
        res_flbl  = Font(bold=True, color=COR_RESUMO_TTL, name="Calibri", size=10)
        res_fval  = Font(bold=True, name="Calibri", size=10)

        def _resumo(label, value):
            ws.append([label, value] + [""] * (len(col_names) - 2))
            rn = ws.max_row
            for cell in ws[rn]:
                cell.fill = res_fill
            ws[rn][0].font      = res_flbl
            ws[rn][0].alignment = aln_left
            ws[rn][1].font      = res_fval
            ws[rn][1].alignment = aln_center
            if isinstance(value, float):
                ws[rn][1].number_format = '#,##0.00'

        _resumo("Divergentes",        contagem.get("Divergente", 0))
        _resumo("Ok",                 contagem.get("Ok", 0))
        _resumo("Somente Fornecedor", contagem.get("Somente Fornecedor", 0))
        _resumo("Conferidos",         contagem.get("Conferido", 0))
        _resumo("Total registros",    sum(contagem.values()))
        _resumo(f"Total Liq. {lbl1}", round(total_liq, 2))

        # ── Largura automática ──
        for col_idx in range(1, len(col_names) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(col_names[col_idx - 1]))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

        caminho = os.path.join(tempfile.gettempdir(), "conciliacao.xlsx")
        wb.save(caminho)
        return caminho

    # ── Pipeline completo ──

    def processar_arquivos(self, caminho1: str, caminho2: str):
        """Lê, agrupa, concilia e gera XLSX. Retorna (resumo, resultado, lbl1, lbl2, xlsx_path)."""
        df1, ext1 = self.ler(caminho1)
        df2, ext2 = self.ler(caminho2)

        lbl1, lbl2 = self.rotulo(ext1), self.rotulo(ext2)
        # Se os dois labels ficaram iguais, diferenciar
        if lbl1 == lbl2:
            lbl1 = "Fornecedor"
            lbl2 = "Sistema Wintour"
        g1 = self.agrupar(df1, ext1)
        g2 = self.agrupar(df2, ext2)

        resultado = self.conciliar(g1, g2, lbl1, lbl2, ext1, ext2)
        xlsx_path = self.gerar_xlsx(resultado, lbl1, lbl2)

        # Resumo por localizador único (não por linha/pax)
        _PIOR = {"Divergente": 0, "Somente Fornecedor": 1, "Conferido": 2, "Ok": 3, "Somente Wintour": 4}
        status_por_loc = {}
        for r in resultado:
            loc = r["loc"]
            if loc not in status_por_loc or _PIOR.get(r["status"], 9) < _PIOR.get(status_por_loc[loc], 9):
                status_por_loc[loc] = r["status"]

        resumo = {
            "lbl1": lbl1,
            "lbl2": lbl2,
            "locs_1": len(g1),
            "locs_2": len(g2),
            "ok":               sum(1 for s in status_por_loc.values() if s == "Ok"),
            "divergentes":      sum(1 for s in status_por_loc.values() if s == "Divergente"),
            "conferidos":       sum(1 for s in status_por_loc.values() if s == "Conferido"),
            "somente_fornecedor": sum(1 for s in status_por_loc.values() if s == "Somente Fornecedor"),
            "somente_wintour":  sum(1 for s in status_por_loc.values() if s == "Somente Wintour"),
        }

        return resumo, resultado, lbl1, lbl2, xlsx_path
